# Get top 30 projects from https://www.defined.fi/ and then fetch top 100 traders from BirdEye API, then export excel file from the result.

import contextlib
import os
import json
import requests
from time import sleep
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook

# Load environment variables
load_dotenv()
birdeye_api = os.getenv("BIRDEYE_API_KEY")

# Constants
HEADLESS_OPTIONS = ["--headless", "--disable-gpu"]
URL_TEMPLATE = (
    "https://public-api.birdeye.so/defi/v2/tokens/top_traders?"
    "address={contract_address}&time_frame=24h&sort_type=desc"
    "&sort_by=volume&offset={offset}&limit=10"
)
EXCEL_FILE = "output.xlsx"
BASE_URL = "https://www.defined.fi/"
TOKENS_URL = (
    f"{BASE_URL}tokens/discover?createdAt=week1&rankingBy=volume"
    "&rankingDirection=DESC&network=sol"
)


def setup_driver():
    chrome_options = Options()
    for option in HEADLESS_OPTIONS:
        chrome_options.add_argument(option)
    return webdriver.Chrome(options=chrome_options)


def find_elements(driver, by, value):
    while True:
        with contextlib.suppress(Exception):
            if elements := driver.find_elements(by, value):
                return elements
        sleep(0.1)


def set_excel():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Top Projects"
    headers = ["Token Name", "Contract Address", "Pair Address"]
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column).value = header
    return wb, sheet


def fetch_token_data(item):
    return {
        "token_name": item.find_element(By.CLASS_NAME, "css-i26l22").text,
        "chain_name": item.find_element(By.CLASS_NAME, "css-i8j6jy")
        .get_attribute("href")
        .replace(BASE_URL, "")
        .split("/")[0],
        "pair_address": item.find_element(By.CLASS_NAME, "css-i8j6jy")
        .get_attribute("href")
        .replace(BASE_URL, "")
        .split("/")[1]
        .split("?")[0],
        "contract_address": item.find_element(By.CLASS_NAME, "css-626yaa")
        .get_attribute("href")
        .split("/")[-1],
        "dex_name": item.find_element(By.CLASS_NAME, "css-1wgwepu").get_attribute(
            "aria-label"
        ),
    }


def get_contract_addresses(driver, wb, sheet):
    output_contract_addresses = []
    start_row = 2

    try:
        driver.get(TOKENS_URL)
        token_rows = find_elements(
            driver, By.CSS_SELECTOR, "div[data-sentry-component='TokenRow']"
        )

        for i, item in enumerate(token_rows[:30], start=1):
            print(f"----------- {i} -----------")
            data = fetch_token_data(item)
            output_contract_addresses.append(data)

            print(
                " / ".join(
                    [
                        data["token_name"],
                        data["chain_name"],
                        data["pair_address"],
                        data["contract_address"],
                        data["dex_name"],
                    ]
                )
            )

            for col, key in enumerate(
                ["token_name", "contract_address", "pair_address"], start=1
            ):
                sheet.cell(row=start_row, column=col).value = data[key]
            start_row += 1

        wb.save(EXCEL_FILE)

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        driver.quit()

    with open("contract_address_list.json", "w", encoding="utf-8") as data_file:
        json.dump(output_contract_addresses, data_file, indent=4)

    return output_contract_addresses


def get_top_trader_address(contract_addresses):
    headers = {
        "accept": "application/json",
        "x-chain": "solana",
        "X-API-KEY": birdeye_api,
    }
    os.makedirs("./top_trader", exist_ok=True)

    for item in contract_addresses:
        wallet_address_list = []

        for offset in range(0, 100, 10):
            url = URL_TEMPLATE.format(
                contract_address=item["contract_address"], offset=offset
            )
            response = requests.get(url, headers=headers)
            data = response.json().get("data", {}).get("items", [])
            wallet_address_list.extend(data)

        with open(f"./top_trader/{item['token_name']}.json", "w") as wallet_data:
            json.dump(wallet_address_list, wallet_data, indent=4)


def append_trader_data_to_excel(contract_addresses):
    workbook = load_workbook(EXCEL_FILE)

    for item in contract_addresses:
        new_sheet = workbook.create_sheet(title=item["token_name"])
        file_path = f"./top_trader/{item['token_name']}.json"

        with open(file_path, "r") as trader_file:
            json_data = json.load(trader_file)

        for index, value in enumerate(json_data):
            new_sheet.cell(row=index + 1, column=1).value = value["owner"]

    workbook.save(EXCEL_FILE)


def main():
    wb, sheet = set_excel()
    driver = setup_driver()
    contract_addresses = get_contract_addresses(driver, wb, sheet)

    get_top_trader_address(contract_addresses)
    append_trader_data_to_excel(contract_addresses)


if __name__ == "__main__":
    main()
